"""Automate copying exposure tables from fund files into a template workbook.

The script is driven by a YAML config (see ``config/fund_profiles.yaml``).
Each fund profile defines how to locate the exposure table in the raw file and
where to paste it into the template.

Example usage:
    python src/auto_copy.py --raw path/to/raw.xlsx \
        --template path/to/template.xlsx \
        --fund berry_street --output output.xlsx

The script preserves cell values only (no formatting). It can clear a target
block before pasting and can skip fully blank rows in the source.
"""
from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Optional

import openpyxl
import yaml


@dataclass
class SourceConfig:
    sheet: str
    range: Optional[str] = None
    start_cell: Optional[str] = None
    stop_at_blank_rows: int = 2
    max_rows: int = 200
    max_cols: int = 50


@dataclass
class TargetConfig:
    sheet: str
    start_cell: str
    clear_rows: int = 0
    clear_cols: int = 0


@dataclass
class FundProfile:
    name: str
    source: SourceConfig
    target: TargetConfig


class ConfigError(Exception):
    """Raised when the YAML config is invalid."""


class ProfileRegistry:
    def __init__(self, config_path: Path) -> None:
        self._profiles = self._load_profiles(config_path)

    def _load_profiles(self, path: Path) -> dict[str, FundProfile]:
        with path.open("r", encoding="utf-8") as fh:
            data = yaml.safe_load(fh) or {}

        funds = data.get("funds", {})
        profiles: dict[str, FundProfile] = {}
        for key, val in funds.items():
            try:
                source_cfg = val["source"]
                target_cfg = val["target"]
            except KeyError as exc:
                raise ConfigError(f"Profile '{key}' is missing required section: {exc}") from exc

            has_range = bool(source_cfg.get("range"))
            has_start = bool(source_cfg.get("start_cell"))
            if has_range and has_start:
                raise ConfigError(
                    f"Profile '{key}' should define either 'range' or 'start_cell', not both"
                )
            if not has_range and not has_start:
                raise ConfigError(
                    f"Profile '{key}' must define either 'range' or 'start_cell' in source config"
                )

            profile = FundProfile(
                name=key,
                source=SourceConfig(
                    sheet=source_cfg["sheet"],
                    range=source_cfg.get("range"),
                    start_cell=source_cfg.get("start_cell"),
                    stop_at_blank_rows=source_cfg.get("stop_at_blank_rows", 2),
                    max_rows=source_cfg.get("max_rows", 200),
                    max_cols=source_cfg.get("max_cols", 50),
                ),
                target=TargetConfig(
                    sheet=target_cfg["sheet"],
                    start_cell=target_cfg["start_cell"],
                    clear_rows=target_cfg.get("clear_rows", 0),
                    clear_cols=target_cfg.get("clear_cols", 0),
                ),
            )
            profiles[key] = profile
        if not profiles:
            raise ConfigError("No fund profiles found in config file")
        return profiles

    def get(self, fund_name: str) -> FundProfile:
        try:
            return self._profiles[fund_name]
        except KeyError as exc:
            raise ConfigError(f"Fund profile '{fund_name}' not found") from exc

    @property
    def available(self) -> Iterable[str]:
        return self._profiles.keys()


def _strip_blank_rows(rows: List[List[object]], allowed_consecutive: int) -> List[List[object]]:
    cleaned: List[List[object]] = []
    blank_run = 0
    for row in rows:
        is_blank = all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row)
        if is_blank:
            blank_run += 1
            if blank_run > allowed_consecutive:
                break
            continue
        blank_run = 0
        cleaned.append(row)
    return cleaned


def _trim_blank_cols(rows: List[List[object]]) -> List[List[object]]:
    if not rows:
        return rows

    last_non_blank_idx = -1
    for row in rows:
        for idx, cell in enumerate(row):
            if cell not in (None, ""):
                last_non_blank_idx = max(last_non_blank_idx, idx)

    if last_non_blank_idx == -1:
        return []

    trim_to = last_non_blank_idx + 1
    return [row[:trim_to] for row in rows]


def _read_range(ws, range_str: str) -> List[List[object]]:
    return [[cell.value for cell in row] for row in ws[range_str]]


def _read_dynamic(ws, start_cell: str, max_rows: int, max_cols: int) -> List[List[object]]:
    start = ws[start_cell]
    values = []
    for r in range(start.row, start.row + max_rows):
        row_vals = []
        for c in range(start.column, start.column + max_cols):
            row_vals.append(ws.cell(row=r, column=c).value)
        values.append(row_vals)
    return values


def _write_table(ws, start_cell: str, table: List[List[object]]) -> None:
    start = ws[start_cell]
    for r_offset, row in enumerate(table):
        for c_offset, value in enumerate(row):
            ws.cell(row=start.row + r_offset, column=start.column + c_offset, value=value)


def _clear_block(ws, start_cell: str, rows: int, cols: int) -> None:
    if rows <= 0 or cols <= 0:
        return
    start = ws[start_cell]
    for r in range(rows):
        for c in range(cols):
            ws.cell(row=start.row + r, column=start.column + c, value=None)


def _load_table(raw_path: Path, profile: FundProfile) -> List[List[object]]:
    raw_wb = openpyxl.load_workbook(raw_path, data_only=True)
    try:
        raw_ws = raw_wb[profile.source.sheet]
    except KeyError as exc:
        raise ConfigError(f"Sheet '{profile.source.sheet}' not found in raw file") from exc

    if profile.source.range:
        table = _read_range(raw_ws, profile.source.range)
    else:
        table = _read_dynamic(
            raw_ws,
            profile.source.start_cell,
            profile.source.max_rows,
            profile.source.max_cols,
        )

    table = _strip_blank_rows(table, profile.source.stop_at_blank_rows)
    table = _trim_blank_cols(table)
    return table


def copy_exposure(raw_path: Path, template_path: Path, output_path: Path, profile: FundProfile) -> None:
    table = _load_table(raw_path, profile)
    template_wb = openpyxl.load_workbook(template_path)

    try:
        target_ws = template_wb[profile.target.sheet]
    except KeyError as exc:
        raise ConfigError(f"Sheet '{profile.target.sheet}' not found in template") from exc

    _clear_block(
        target_ws,
        start_cell=profile.target.start_cell,
        rows=profile.target.clear_rows,
        cols=profile.target.clear_cols,
    )
    _write_table(target_ws, start_cell=profile.target.start_cell, table=table)

    template_wb.save(output_path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--raw", required=True, type=Path, help="Path to the raw fund Excel file")
    parser.add_argument("--template", required=True, type=Path, help="Path to the template Excel file")
    parser.add_argument("--fund", required=True, help="Fund profile name to use")
    parser.add_argument("--config", default=Path("config/fund_profiles.yaml"), type=Path, help="YAML config path")
    parser.add_argument("--output", required=True, type=Path, help="Where to write the populated template")
    parser.add_argument(
        "--list-funds",
        action="store_true",
        help="List available fund profiles from the config and exit",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and preview the table without writing the output file",
    )
    return parser


def main(argv: Optional[List[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    registry = ProfileRegistry(args.config)

    if args.list_funds:
        for name in sorted(registry.available):
            print(name)
        return 0

    profile = registry.get(args.fund)

    table = _load_table(args.raw, profile)

    if args.dry_run:
        preview_rows = table[:5]
        print(f"Fund: {profile.name}")
        print(f"Rows: {len(table)} | Cols: {len(preview_rows[0]) if preview_rows else 0}")
        for row in preview_rows:
            print(row)
        return 0

    # Reuse copy_exposure for the actual write path to keep clearing/target logic centralized.
    copy_exposure(args.raw, args.template, args.output, profile)
    print(f"Copied exposure for {profile.name} -> {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
